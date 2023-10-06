import pandas as pd
import numpy as np
import openpyxl

#入参为list，返回排序后的中位数
def get_mid(y):
    x = y
    length = len(x)
    x.sort()
    z=length // 2
    n = x[z]    # 中位数取靠中间右侧的
    return n

def get_one(b):     #去除最值后 求中值
    a = b
    if len(a) == 1:
        return a[0]
    elif len(a) == 2:       #两个数据的情况，取之后的那个
        return a[1]
    else:
        a.remove(np.min(a))
        a.remove(np.max(a))
        return get_mid(a)

# 定义要读写的excel文件路径
# file_path1 = 'data_source/out_all_ages_new_quxia1.0.xlsx' # 一个名字有多行 数值表现形式为中心值 误差两列的那个 现在数据还没去杂质
# file_path1 = 'data_source/out_all_ages_new_tag_quxia1.xlsx' # 一个名字有多行 数值表现形式为中心值 误差两列的那个  AE未处理的时候
file_path1 = 'final_date/3_1_quxia1.0.xlsx' # 一个名字有多行 数值表现形式为中心值 误差两列 已经阈值=1去除瑕疵了


# print_log = open("out/5w_to_one_new0915.txt",'w')
# print_log = open("out/new_tag_quxia_one.txt",'w')
print_log = open("final_date/3_2.txt",'w')

write_add = openpyxl.load_workbook(file_path1)    # 打开原始Excel文件,用openpyxl模块进行写操作

write_sheet = write_add['Sheet1']     #当前工作sheet
wz_all_num = write_sheet.max_row - 1    #当前sheet行数
# print(wz_all_num)
kc_name_list = []
kc_ages_list = []
# age_err_list = []
count_name = 0
len_ages = 0
for row_value in range(2,wz_all_num + 2) :   #遍历每一行,先不考虑误差
    here_kc = write_sheet.cell(row=row_value,column=1).value
    here_d = write_sheet.cell(row=row_value,column=3).value
    # here_err = write_sheet.cell(row=row_value,column=5).value
    last_kc =  write_sheet.cell(row=row_value-1,column=1).value

    # print(here_kc,here_d,file=print_log)
    if len(kc_name_list) == 0:
        last_kc = here_kc
    if here_kc not in kc_name_list:
        kc_name_list.append(here_kc)
        # count_name +=1
        # len_ages += len(kc_ages_list)
        if len(kc_ages_list)>0:

            # print(last_kc,kc_ages_list,age_err_list,file=print_log)
            ff_age = get_one(kc_ages_list)
            # ff_err = age_err_list[kc_ages_list.index(ff_age)]
            print(last_kc+','+str(ff_age),file=print_log)


        kc_ages_list = []
        # age_err_list = []
        kc_ages_list.append(float(here_d))
        # age_err_list.append(here_err)
    else :
        kc_ages_list.append(float(here_d))
        # age_err_list.append(here_err)

final_kc = write_sheet.cell(row=wz_all_num+1,column=1).value
# print(final_kc,kc_ages_list,age_err_list,file=print_log)      #最后一个矿床
# len_ages += len(kc_ages_list)
ff_age1 = get_one(kc_ages_list)
print(last_kc+','+str(ff_age1),file=print_log)
print(kc_name_list,file=print_log)

print(len(kc_name_list))