# -*- coding: UTF-8 -*-
import re
import json
import pandas as pd
import openpyxl

# 0.2 入参为一个句子，调用正则表达式，返回句子的age信息
def extract_V2(str1):
    pattern = r'\b(?:from\s+)?\d+(?:\.\d+)?(?:\s*±\s*\d+(?:\.\d+)?)?(?:\s+(?:to|similar\sto|Ma\sto)\s+\d+(?:\.\d+)?|\s*-\s*\d+(?:\.\d+)?)?\s*Ma\b'
    matches = re.findall(pattern, str1)
    return matches

# 0.1 excel 矿床名一列转为列表返回 [11, 45, 14]形式
def out_kc():
    file_path = 'data_source/kc_name.xlsx'  # 替换为你的 Excel 文件路径 存放的给定的矿床名称
    column_name = 'kc_NAME'  # 替换为你要读取的列名

    try:
        df = pd.read_excel(file_path, engine='openpyxl')
        
        if column_name in df.columns:
            column_data = df[column_name].tolist()
            # print(column_data)
            return column_data
        else:
            print(f"Column '{column_name}' not found in the Excel file.")
    except FileNotFoundError:
        print(f"File not found: {file_path}")
    except Exception as e:
        print(f"An error occurred: {str(e)}")

# 0 注释 传入一个文件路径，按行抽取age信息
def run_all(file_path): #file_path 是指向清洗重复句子之后的txt文件吧 
    all_dict = {} # 创建一个空字典 键值对
    list_kc_name = out_kc() # 0.1
    # print(len(list_kc_name))
    print_log = open("out/0911.txt", 'w',encoding='UTF-8')
    try:
        with open(file_path, 'r',encoding='UTF-8') as file2:
            for line in file2:
                # 使用strip()方法去除行末尾的换行符
                cleaned_line = line.strip()
                for kc in list_kc_name:
                    kc1 = kc.strip() #
                    # 9.8 修改 增加下一行 解决名字(铜山 铜山岭)抽取错误 
                    kc1 = kc1 + ' '
                    if kc1 in cleaned_line:
                        this_age_list = extract_V2(cleaned_line) #0.2 调用正则表达式函数抽取
                        # print(this_age_list)
                        if len(this_age_list)>0:
                            # this_kc_index = cleaned_line.find(kc1)
                            # print(kc1,this_kc_index)
                            # this_distance = 9999
                            # nearest_age = ''
                            # for age1 in this_age_list:
                            #     this_age_index = cleaned_line.find(age1)
                            #     if (abs(this_kc_index-this_age_index)<this_distance):
                            #         this_distance = abs(this_kc_index-this_age_index)
                            #         print(kc1,this_kc_index,this_distance)
                            #         nearest_age = age1
                            #         this_age_list.remove(age1)


                            # all_dict.update({cleaned_line+'@@@@@'+kc1:this_age_list})
                            print(cleaned_line+'￥'+kc1+'￥'+str(this_age_list),file=print_log) # 输出三列 句子 矿床 年龄(多形式)
                    else:
                        continue


    except FileNotFoundError:
        print(f"File not found: {file_path}")
    except Exception as e:
        print(f"An error occurred: {str(e)}")    
    return all_dict

# 1.1 
def is_number(s):
    try:  # 如果能运行float(s)语句，返回True（字符串s是浮点数）
        float(s)
        return True
    except ValueError:  # ValueError为Python的一种标准异常，表示"传入无效的参数"
        pass  # 如果引发了ValueError这种异常，不做任何事情（pass：不做任何事情，一般用做占位语句）
    try:
        import unicodedata  # 处理ASCii码的包
        unicodedata.numeric(s)  # 把一个表示数字的字符串转换为浮点数返回的函数
        return True
    except (TypeError, ValueError):
        pass
    return False

# 1 把excel文件中的age信息全部归到一起 输出日志中 由一个单元格多种形式数据拆分为多行 中心值±+误差两列的形式
# 下一步手动excel分割钱币符号 输出out_all_ages.xlsx文件 带有中心值误差值的表格 仍有多行为同一个矿床名字的那种 下一步要变成一行的（去最大最小 取中位数等措施）
def age_out(file_path):
    op_excel = openpyxl.load_workbook(file_path)
    op_sheet = op_excel['Sheet1']     #当前工作sheet
    count_row = op_sheet.max_row - 1    #当前sheet行数
    print_log = open("result/all_v2.log",'w')
    # print(count_row)
    for row_index in range(2,count_row + 2) :   #遍历每一行
        if op_sheet.cell(row=row_index,column=3).value != None :
            this_kc = op_sheet.cell(row=row_index,column=2).value
            this_sentence = op_sheet.cell(row=row_index,column=1).value
            original_age = op_sheet.cell(row=row_index,column=3).value
            
            list_age1 = eval(original_age) #识别为列表类型的age
            # print(this_kc,list_age1)
            if (len(list_age1)>0):
                for age1 in list_age1:
                    # print(age1)
                    if age1 == 'Ma' or age1 == ' Ma' or age1 ==' Ma ':
                        continue
                    else:
                        # age1 = age1.replace(' Ma', '')    #去掉Ma后的时间
                        # # print(kc, wz_num,row_value-1,here_ma,age1,file = print_log)
                        age1 = age1.replace('-',' - ')
                        age_str_list = age1.split(' ')
                        # print(age_str_list)
                        count_digit = 0
                        out_here_age_num = []
                        for here_split in age_str_list:
                            here_split = here_split.strip()
                            # 1.1 is_number 函数 判断字符串中有几个数字出现 匹配不同策略
                            if is_number(here_split):
                                count_digit += 1    #看字符串中有几个数字
                                out_here_age_num.append(here_split)
                        # print(count_digit, file = print_log)
                        if count_digit == 0: # 没有数字(年龄)的
                            continue
                        elif count_digit == 1: #只有一个年龄的 直接填充
                            # print('1',out_here_age_num[0],'None',file=print_log)
                            # 2023.9.8 添加下一行 大于4500的扔掉
                            if float(out_here_age_num[0]) < 4500:
                                print(this_kc+'￥'+this_sentence+'￥'+out_here_age_num[0]+'￥'+'None',file=print_log) # 矿床名字 句子 中心值 误差值
                        else :

                            if age1.find('±')!=-1  :
                                mid_age1 = out_here_age_num[0]
                                err_age1 = out_here_age_num[1]
                                # 2023.9.8 添加下一行 大于4500的扔掉
                                if float(mid_age1) <4500:
                                    print(this_kc+'￥'+this_sentence+'￥'+mid_age1+'￥'+err_age1,file=print_log) # 矿床名字 句子 中心值 误差值
                            elif age1.find('to')!= -1 or age1.find('-')!= -1:
                                mid_age2 = (float(out_here_age_num[0])+float(out_here_age_num[1]))/2
                                err_age2 = round(abs(float(out_here_age_num[0])-mid_age2),2)
                                # 2023.9.8 添加下一行 大于4500的扔掉
                                if mid_age2 < 4500:
                                    print(this_kc+'￥'+this_sentence+'￥'+str(mid_age2)+'￥'+str(err_age2),file=print_log) # 矿床名字 句子 中心值 误差值
    print_log.close()



if __name__ == '__main__':
    filetxt_path = 'out/all_Ma_out.txt' 
    # file1_path = 'data_source/all_5w_ages.xlsx' 
    run_all(filetxt_path) # 0 
    # run_all(file1_path) # 0 
    # age_out(file1_path)   # 1 

    
