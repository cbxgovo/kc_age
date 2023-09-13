'''

2023.09.08
multxt_to_onetxt1()函数读取带有重复句子的txt,保留不重复的重新输出txt文件
remove_error()函数将两列 根据同名数据的正态分布原则去除误差值并保留可信度较高的行,输出到Excel数据

2023.09.09
把师兄的2、3函数原来的基础上修改跑通了
结合09.08的进度 并且在to_one函数先不变的情况下 最终数据已经整理出来了
下一步想办法把去掉瑕疵数据后的多个数据归一 找找用什么办法或者算法合理

'''

import re
import json
import pandas as pd
import openpyxl


# 直接函数方法 低行数测试
def multxt_to_onetxt2():
    fi = open('out/test1.txt', 'r')  # 打开需要处理的test.txt。
    txt = fi.readlines()
    with open('out/test1_out.txt', 'a') as f:  # 创建处理去重复后的结果保存文档，防止找不到文件出错
        f.close()
    for w in txt:
        fi2 = open('out/test1_out.txt', 'r')
        txt2 = fi2.readlines()
        with open('out/test1_out.txt', 'a') as f:  # 打开目标文件开始写入
            if w not in txt2:  # 如果从源文档中读取的内容不在目标文档中则写入，否则跳过，实现去除重复功能！
                f.write(w)
            else:
                print("已去除重复-->"+w)
            f.close()
    fi.close()


# * 1 直接函数方法 去重函数 去除txt文件中的重复句子重新保存到txt文件中
def multxt_to_onetxt1():
    # fi = open('out/all_Ma.txt', 'r', encoding='UTF-8')  # 打开需要处理的test.txt。
    # txt = fi.readlines()
    # with open('out/all_Ma_out.txt', 'a', encoding='UTF-8') as f:  # 创建处理去重复后的结果保存文档，防止找不到文件出错
    #     f.close()
    # for w in txt:
    #     fi2 = open('out/all_Ma_out.txt', 'r', encoding='UTF-8')
    #     txt2 = fi2.readlines()
    #     with open('out/all_Ma_out.txt', 'a', encoding='UTF-8') as f:  # 打开目标文件开始写入
    #         if w not in txt2:  # 如果从源文档中读取的内容不在目标文档中则写入，否则跳过，实现去除重复功能！
    #             f.write(w)
    #         else:
    #             print("已去除重复-->"+w)
    #         f.close()
    # fi.close()
    print_log = open("out/all_Ma_out.txt", 'w',encoding='UTF-8')
    fi = open('out/all_Ma.txt', 'r', encoding='UTF-8')  # 打开需要处理的test.txt。

    txt = fi.readlines()
    print(len(txt))
    list1 = []
    for i in txt:
        if i not in list1 :
            list1.append(i)
    print(len(list1),file=print_log)
    for i in list1:
        print(i,file=print_log)


# 2.注释 传入一个文件路径，按行抽取age信息
def run_all(file_path):  # file_path 是指向清洗重复句子之后的txt文件吧
    all_dict = {}  # 创建一个空字典 键值对
    list_kc_name = out_kc()  # 0.1

    # 全局变量 用于控制书写excel的行数
    global row_num
    row_num = 2
    # 创建一个新的Excel工作簿 覆盖all_5w_ages.xlsx
    workbook = openpyxl.Workbook()

    # 创建一个新的工作表
    sheet = workbook.active
    # 设置表头
    sheet['A1'] = 'Sentences'
    sheet['B1'] = 'kc_name'
    sheet['C1'] = 'age'

    # print(len(list_kc_name))
    try:
        with open(file_path, 'r', encoding='UTF-8') as file2:
            for line in file2:
                # 使用strip()方法去除行末尾的换行符
                cleaned_line = line.strip()
                for kc in list_kc_name:
                    kc1 = kc.strip()
                    # 9.8 修改 增加下一行 解决名字(铜山 铜山岭)抽取错误
                    kc1 = kc1 + ' '
                    if kc1 in cleaned_line:
                        this_age_list = extract_V2(cleaned_line)  # 0.2 调用正则表达式函数抽取
                        # print(this_age_list)
                        if len(this_age_list) > 0:
                            # this_kc_index = cleaned_line.find(kc1)
                            # print(kc1,this_kc_index)
                            # this_distance = 9999
                            # nearest_age = ''
                            # for age1 in this_age_list:
                            #     this_age_index = cleaned_line.find(age1)
                            #     if (abs(this_kc_index-this_age_index)<this_distance):
                            #         this_distance = abs(this_kc_index-this_age_index)
                            #         print(kc1,this_kc_index,this_distance)
                            #         nearest_age = age1
                            #         this_age_list.remove(age1)

                            # all_dict.update({cleaned_line+'@@@@@'+kc1:this_age_list})
                            # 输出三列 句子 矿床 年龄(多形式)
                            # print(cleaned_line+'￥'+kc1+'￥'+str(this_age_list)) # 控制台打印输出 - 替换直接输入excel表格中

                            # 遍历数据列表，将数据写入Excel表格
                            # row_num = 2  # 从第二行开始写入数据，因为第一行是表头 得设置成全局变量
                            sheet.cell(row=row_num, column=1,value=cleaned_line)
                            sheet.cell(row=row_num, column=2, value=kc1)
                            sheet.cell(row=row_num, column=3,value=str(this_age_list))
                            row_num += 1
                            # 保存Excel文件
                            print("已经保存", row_num-2, '行')
                    else:
                        continue
            # 保存循环之后的excel表格
            workbook.save('data_source/all_5w_ages_out.xlsx')

    except FileNotFoundError:
        print(f"File not found: {file_path}")
    except Exception as e:
        print(f"An error occurred: {str(e)}")
    return all_dict


# 2.1 excel 矿床名一列转为列表返回 [11, 45, 14]形式
def out_kc():
    file_path = 'data_source/kc_name.xlsx'  # 替换为你的 Excel 文件路径 存放的给定的矿床名称
    column_name = 'kc_NAME'  # 替换为你要读取的列名

    try:
        # 使用pandas库来读取或写入Excel文件，你可以指定使用openpyxl引擎
        df = pd.read_excel(file_path, engine='openpyxl')

        if column_name in df.columns:
            column_data = df[column_name].tolist()
            # print(column_data)
            return column_data
        else:
            print(f"Column '{column_name}' not found in the Excel file.")
    except FileNotFoundError:
        print(f"File not found: {file_path}")
    except Exception as e:
        print(f"An error occurred: {str(e)}")
    

# 2.2 入参为一个句子，调用正则表达式，返回句子的age信息
def extract_V2(str1):
    pattern = r'\b(?:from\s+)?\d+(?:\.\d+)?(?:\s*±\s*\d+(?:\.\d+)?)?(?:\s+(?:to|similar\sto|Ma\sto)\s+\d+(?:\.\d+)?|\s*-\s*\d+(?:\.\d+)?)?\s*Ma\b'
    matches = re.findall(pattern, str1)
    return matches


# 3 把excel文件中的age信息全部归到一起 输出日志中 由一个单元格多种形式数据拆分为多行 中心值±+误差两列的形式
# 下一步手动excel分割钱币符号 输出out_all_ages.xlsx文件 带有中心值误差值的表格 仍有多行为同一个矿床名字的那种 下一步要变成一行的（去最大最小 取中位数等措施）
def age_out(file_path):
    # 全局变量 用于控制书写excel的行数
    global row_num1
    row_num1 = 2
    # 创建一个新的Excel工作簿 覆盖all_5w_ages.xlsx
    workbook = openpyxl.Workbook()

    # 创建一个新的工作表
    sheet = workbook.active
    # 设置表头
    sheet['A1'] = 'kc'
    sheet['B1'] = 'cleanen_line'
    sheet['C1'] = 'age'
    sheet['D1'] = 'err'

    op_excel = openpyxl.load_workbook(file_path)
    op_sheet = op_excel['Sheet1']  # 当前工作sheet
    count_row = op_sheet.max_row - 1  # 当前sheet行数
    print_log = open("result/all_v2.log", 'w')
    # print(count_row)
    for row_index in range(2, count_row + 2):  # 遍历每一行
        if op_sheet.cell(row=row_index, column=3).value != None:
            this_kc = op_sheet.cell(row=row_index, column=2).value
            this_sentence = op_sheet.cell(row=row_index, column=1).value
            original_age = op_sheet.cell(row=row_index, column=3).value

            list_age1 = eval(original_age)  # 识别为列表类型的age
            # print(this_kc,list_age1)
            if (len(list_age1) > 0):
                for age1 in list_age1:
                    # print(age1)
                    if age1 == 'Ma' or age1 == ' Ma' or age1 == ' Ma ':
                        continue
                    else:
                        # age1 = age1.replace(' Ma', '')    #去掉Ma后的时间
                        # # print(kc, wz_num,row_value-1,here_ma,age1,file = print_log)
                        age1 = age1.replace('-', ' - ')
                        age_str_list = age1.split(' ')
                        # print(age_str_list)
                        count_digit = 0
                        out_here_age_num = []
                        for here_split in age_str_list:
                            here_split = here_split.strip()
                            # 1.1 is_number 函数 判断字符串中有几个数字出现 匹配不同策略
                            if is_number(here_split):
                                count_digit += 1  # 看字符串中有几个数字
                                out_here_age_num.append(here_split)
                        # print(count_digit, file = print_log)
                        if count_digit == 0:  # 没有数字(年龄)的
                            continue
                        elif count_digit == 1:  # 只有一个年龄的 直接填充
                            # print('1',out_here_age_num[0],'None',file=print_log)
                            # 2023.9.8 添加下一行 大于4500的扔掉
                            if float(out_here_age_num[0]) < 4500:
                                # 矿床名字 句子 中心值 误差值
                                # print(this_kc+'￥'+this_sentence+'￥' +out_here_age_num[0]+'￥'+'None', file=print_log) # 用excel直接书写代替该行
                                sheet.cell(row=row_num1, column=1,value=this_kc)
                                sheet.cell(row=row_num1, column=2, value=this_sentence)
                                sheet.cell(row=row_num1, column=3,value=out_here_age_num[0])
                                sheet.cell(row=row_num1, column=4,value='None')
                                row_num1 += 1
                                print("已经保存", row_num1-2, '行')
                        else:

                            if age1.find('±') != -1:
                                mid_age1 = out_here_age_num[0]
                                err_age1 = out_here_age_num[1]
                                # 2023.9.8 添加下一行 大于4500的扔掉
                                if float(mid_age1) < 4500:
                                    # 矿床名字 句子 中心值 误差值
                                    # print(this_kc+'￥'+this_sentence+'￥' +mid_age1+'￥'+err_age1, file=print_log)
                                    sheet.cell(row=row_num1, column=1,value=this_kc)
                                    sheet.cell(row=row_num1, column=2, value=this_sentence)
                                    sheet.cell(row=row_num1, column=3,value=mid_age1)
                                    sheet.cell(row=row_num1, column=4,value=err_age1)
                                    row_num1 += 1
                                    print("已经保存", row_num1-2, '行')
                            elif age1.find('to') != -1 or age1.find('-') != -1:
                                mid_age2 = (
                                    float(out_here_age_num[0])+float(out_here_age_num[1]))/2
                                err_age2 = round(abs(float(out_here_age_num[0])-mid_age2), 2)
                                # 2023.9.8 添加下一行 大于4500的扔掉
                                if mid_age2 < 4500:
                                    # 矿床名字 句子 中心值 误差值
                                    # print(this_kc+'￥'+this_sentence+'￥'+str(mid_age2)+'￥'+str(err_age2), file=print_log)
                                    sheet.cell(row=row_num1, column=1,value=this_kc)
                                    sheet.cell(row=row_num1, column=2, value=this_sentence)
                                    sheet.cell(row=row_num1, column=3,value=mid_age2)
                                    sheet.cell(row=row_num1, column=4,value=err_age2)
                                    row_num1 += 1
                                    print("已经保存", row_num1-2, '行')
    # print_log.close()
    # 保存循环之后的excel表格
    workbook.save('data_source/out_all_ages_new.xlsx')


# 3.1
def is_number(s):
    try:  # 如果能运行float(s)语句，返回True（字符串s是浮点数）
        float(s)
        return True
    except ValueError:  # ValueError为Python的一种标准异常，表示"传入无效的参数"
        pass  # 如果引发了ValueError这种异常，不做任何事情（pass：不做任何事情，一般用做占位语句）
    try:
        import unicodedata  # 处理ASCii码的包
        unicodedata.numeric(s)  # 把一个表示数字的字符串转换为浮点数返回的函数
        return True
    except (TypeError, ValueError):
        pass
    return False


# 4.根据同名数据的正态分布原则去除误差值并保留可信度较高的行 会丢失单行和两行数据差距较大的行修改
'''
1.读取Excel文件。
2.根据名字分组数据。
3.对每个分组进行正态分布异常值检测，保留可信度较高的数据。
4.将处理后的数据输出到新的Excel文件。
'''
def remove_error0():
    import pandas as pd
    import numpy as np
    from scipy import stats

    # 读取Excel文件
    df = pd.read_excel('data_source/test0911.xlsx')
    # 定义阈值（可以根据需要调整） 越小越严格 剩下的越少 2.0表示使用了一个阈值为2.0的标准差 在正态分布中，大约95.4% 的数据点落在均值的两个标准差范围内
    # 原来  8.4w行
    # 散点图 Aktogai矿床为例子  Altar为例子
    # `threshold = 2.0`   只保留大约均值附近95.4%   剩余8w行
    # `threshold = 1.0`  只保留大约均值附近68.26%   剩余7.3w行
    # `threshold = 0.5`  只保留大约均值附近38.3%    剩余5.6w行  较好
    # `threshold = 0.1`  只保留大约均值附近7.96%    剩余1w行
    threshold = 1.0

    # 创建一个新的DataFrame来存储处理后的数据
    cleaned_df = pd.DataFrame(columns=df.columns)

    # 根据名字分组数据并处理每个分组 数据集group
    for name, group in df.groupby('kc'):  # 根据第一列矿床的名字进行分组
        # 计算每个分组的均值和标准差
        # print(type(group['age']))
        # print(group['age'])
        print(len(group['age']))

        mean = group['age'].mean()  # 得到'age'列的均值 mean变量保存这个均值
        std_dev = group['age'].std()  # 保存'age'列的标准差

        # 使用正态分布的方法去除异常值
        # 计算Z分数 Z分数越大，表示数据点偏离均值越远。
        if(len(group['age'])>2):
            z_scores = np.abs(stats.zscore(group['age']))
            valid_rows = (z_scores < threshold)  # 合理行 定义为偏离小于阈值的部分

            # 保留可信度较高的行
            cleaned_group = group[valid_rows]

            # 将处理后的分组数据添加到新的DataFrame中
            cleaned_df = pd.concat([cleaned_df, cleaned_group])
        else:
            z_scores = np.abs(stats.zscore(group['age']))
            print(type(z_scores))
            valid_rows = ((z_scores< threshold) | (z_scores> threshold) | (z_scores is  threshold) )  # 合理行 
            cleaned_group = group[valid_rows]
            cleaned_df = pd.concat([cleaned_df, cleaned_group])
    # 将处理后的数据保存到新的Excel文件
    cleaned_df.to_excel('data_source/test0911_out.xlsx', index=False)


# 4.根据同名数据的正态分布原则去除误差值并保留可信度较高的行
'''
1.读取Excel文件。
2.根据名字分组数据。
3.对每个分组进行正态分布异常值检测，保留可信度较高的数据。
4.将处理后的数据输出到新的Excel文件。
'''
def remove_error1():
    import pandas as pd
    import numpy as np
    from scipy import stats

    # 读取原始Excel文件
    # df = pd.read_excel('data_source/test0911.xlsx')
    df = pd.read_excel('data_source/out_all_ages_new.xlsx')

    # 按照名字分组
    grouped = df.groupby('kc')

    # 创建一个空的DataFrame来存储结果
    filtered_data = pd.DataFrame(columns=df.columns)

    # `threshold = 2.0`   只保留大约均值附近95.4%  
    # `threshold = 1.0`  只保留大约均值附近68.26%   
    # `threshold = 0.5`  只保留大约均值附近38.3%    
    # `threshold = 0.1`  只保留大约均值附近7.96%   
    threshold = 1.0

    # 遍历每个名字的数据
    for name, group_data in grouped:
        if len(group_data) <= 2:  # 如果数据少于等于2个，直接保留
            filtered_data = pd.concat([filtered_data, group_data], ignore_index=True)
        else:
            # 计算数据的均值和标准差
            mean = np.mean(group_data['age'])
            std_dev = np.std(group_data['age'])
            
            # 使用均值和标准差来过滤数据
            filtered_group_data = group_data[abs(group_data['age'] - mean) <= threshold * std_dev] # threshold 数字越小 保留越少
            filtered_data = pd.concat([filtered_data, filtered_group_data], ignore_index=True)

    # 将结果保存到新的Excel文件
    # filtered_data.to_excel('data_source/test0911_out.xlsx', index=False)
    filtered_data.to_excel('data_source/out_all_ages_new_quxia1.0.xlsx', index=False)
    

def test_mean():
    import pandas as pd

    # 读取原始Excel文件
    df = pd.read_excel('data_source/out_all_ages_new_quxia1.0.xlsx')

    # 使用groupby按照名字分组，并计算每个分组中数字列的平均值
    result_df = df.groupby('kc', as_index=False)['age'].mean() # kc和age分别为第一列和第三列的表格列名

    # 创建一个新的Excel工作簿并将结果写入
    result_df.to_excel('data_source/out_all_ages_new_quxia1.0_mean.xlsx', index=False)





if __name__ == '__main__':
    filetxt_path = 'out/all_Ma_out.txt'
    file1_path = 'data_source/all_5w_ages_new.xlsx'

    # 1.调用去重函数 去除txt文件中的重复句子
    # multxt_to_onetxt1()
    # print("已全部去除重复.")

    # 2.去重后的txt提取到第一个三列的表格 第三列为多种数据格式
    # run_all(filetxt_path)  

    # 3.三列未拆分的表格拆分为多行(中心值和误差两列数据的形式) 矿名仍有重复
    # age_out(file1_path)   # 1 第一版的表格 - > 输出多行表格  中心值 误差两列数据形式 pandas版本问题

    # 4.调用去除误差数据的函数 去除偏差较大的数据 保留较为集中的一部分 正态 
    #   remove_error0()有问题 有的矿床名字下的数据可能距离均值都比较远 都被扔了 还要保留么这种
    #   remove_error1()改进了 没问题了 不丢失
    # remove_error1()
    # print("处理后的数据已经保存到新的Excel文件.")

    # 5.想合适的算法把去掉异常值之后的数据 归一



    # 6.假如取 去除瑕疵之后的 平均值的话
    # test_mean()


