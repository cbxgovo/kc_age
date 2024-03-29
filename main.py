'''

2023.09.08
multxt_to_onetxt1()函数读取带有重复句子的txt,保留不重复的重新输出txt文件
remove_error()函数将两列 根据同名数据的正态分布原则去除误差值并保留可信度较高的行,输出到Excel数据

2023.09.09
把师兄的2、3函数原来的基础上修改跑通了
结合09.08的进度 并且在to_one函数先不变的情况下 最终数据已经整理出来了
下一步想办法把去掉瑕疵数据后的多个数据归一 找找用什么办法或者算法合理

2023.09.20
按照属性1提取的tag弄出来了
将归一数据和美国地调局的数据对比做完了
将按照属性1提取的年龄和美国地调局的数据对比做完了

'''

import re
import json
import pandas as pd
import openpyxl


# 直接函数方法 低行数测试
def multxt_to_onetxt2():
    fi = open('out/test1.txt', 'r')  # 打开需要处理的test.txt。
    txt = fi.readlines()
    with open('out/test1_out.txt', 'a') as f:  # 创建处理去重复后的结果保存文档，防止找不到文件出错
        f.close()
    for w in txt:
        fi2 = open('out/test1_out.txt', 'r')
        txt2 = fi2.readlines()
        with open('out/test1_out.txt', 'a') as f:  # 打开目标文件开始写入
            if w not in txt2:  # 如果从源文档中读取的内容不在目标文档中则写入，否则跳过，实现去除重复功能！
                f.write(w)
            else:
                print("已去除重复-->"+w)
            f.close()
    fi.close()


# * 1 直接函数方法 去重函数 去除txt文件中的重复句子重新保存到txt文件中
def multxt_to_onetxt1():
    fi = open('out/all_Ma.txt', 'r', encoding='UTF-8')  # 打开需要处理的test.txt。
    txt = fi.readlines()
    with open('out/all_Ma_out.txt', 'a', encoding='UTF-8') as f:  # 创建处理去重复后的结果保存文档，防止找不到文件出错
        f.close()
    for w in txt:
        fi2 = open('out/all_Ma_out.txt', 'r', encoding='UTF-8')
        txt2 = fi2.readlines()
        with open('out/all_Ma_out.txt', 'a', encoding='UTF-8') as f:  # 打开目标文件开始写入
            if w not in txt2:  # 如果从源文档中读取的内容不在目标文档中则写入，否则跳过，实现去除重复功能！
                f.write(w)
            else:
                print("已去除重复-->"+w)
            f.close()
    fi.close()
    # print_log = open("out/all_Ma_out.txt", 'w',encoding='UTF-8')
    # fi = open('out/all_Ma.txt', 'r', encoding='UTF-8')  # 打开需要处理的test.txt。

    # txt = fi.readlines()
    # print(len(txt))
    # list1 = []
    # for i in txt:
    #     if i not in list1 :
    #         list1.append(i)
    # print(len(list1),file=print_log)
    # for i in list1:
    #     print(i,file=print_log)


# 2.注释 传入一个文件路径，按行抽取age信息
def run_all(file_path):  # file_path 是指向清洗重复句子之后的txt文件吧
    all_dict = {}  # 创建一个空字典 键值对
    list_kc_name = out_kc()  # 0.1

    # 全局变量 用于控制书写excel的行数
    global row_num
    row_num = 2
    # 创建一个新的Excel工作簿 覆盖all_5w_ages.xlsx
    workbook = openpyxl.Workbook()

    # 创建一个新的工作表
    sheet = workbook.active
    # 设置表头
    sheet['A1'] = 'Sentences'
    sheet['B1'] = 'kc_name'
    sheet['C1'] = 'age'

    # print(len(list_kc_name))
    try:
        with open(file_path, 'r', encoding='UTF-8') as file2:
            for line in file2:
                # 使用strip()方法去除行末尾的换行符
                cleaned_line = line.strip()
                for kc in list_kc_name:
                    kc1 = kc.strip()
                    # 9.8 修改 增加下一行 解决名字(铜山 铜山岭)抽取错误
                    kc1 = kc1 + ' '
                    if kc1 in cleaned_line:
                        this_age_list = extract_V2(cleaned_line)  # 0.2 调用正则表达式函数抽取
                        # print(this_age_list)
                        if len(this_age_list) > 0:
                            # this_kc_index = cleaned_line.find(kc1)
                            # print(kc1,this_kc_index)
                            # this_distance = 9999
                            # nearest_age = ''
                            # for age1 in this_age_list:
                            #     this_age_index = cleaned_line.find(age1)
                            #     if (abs(this_kc_index-this_age_index)<this_distance):
                            #         this_distance = abs(this_kc_index-this_age_index)
                            #         print(kc1,this_kc_index,this_distance)
                            #         nearest_age = age1
                            #         this_age_list.remove(age1)

                            # all_dict.update({cleaned_line+'@@@@@'+kc1:this_age_list})
                            # 输出三列 句子 矿床 年龄(多形式)
                            # print(cleaned_line+'￥'+kc1+'￥'+str(this_age_list)) # 控制台打印输出 - 替换直接输入excel表格中

                            # 遍历数据列表，将数据写入Excel表格
                            # row_num = 2  # 从第二行开始写入数据，因为第一行是表头 得设置成全局变量
                            sheet.cell(row=row_num, column=1,value=cleaned_line)
                            sheet.cell(row=row_num, column=2, value=kc1)
                            sheet.cell(row=row_num, column=3,value=str(this_age_list))
                            row_num += 1
                            # 保存Excel文件
                            print("已经保存", row_num-2, '行')
                    else:
                        continue
            # 保存循环之后的excel表格
            # workbook.save('data_source/all_5w_ages_out.xlsx') # 未去AE之前的
            workbook.save('final_date/1_oneLine.xlsx')

    except FileNotFoundError:
        print(f"File not found: {file_path}")
    except Exception as e:
        print(f"An error occurred: {str(e)}")
    return all_dict


# 2.1 excel 矿床名一列转为列表返回 [11, 45, 14]形式
def out_kc():
    file_path = 'data_source/kc_name.xlsx'  # 替换为你的 Excel 文件路径 存放的给定的矿床名称
    column_name = 'kc_NAME'  # 替换为你要读取的列名

    try:
        # 使用pandas库来读取或写入Excel文件，你可以指定使用openpyxl引擎
        df = pd.read_excel(file_path, engine='openpyxl')

        if column_name in df.columns:
            column_data = df[column_name].tolist()
            # print(column_data)
            return column_data
        else:
            print(f"Column '{column_name}' not found in the Excel file.")
    except FileNotFoundError:
        print(f"File not found: {file_path}")
    except Exception as e:
        print(f"An error occurred: {str(e)}")
    

# 2.2 入参为一个句子，调用正则表达式，返回句子的age信息
def extract_V2(str1):
    pattern = r'\b(?:from\s+)?\d+(?:\.\d+)?(?:\s*±\s*\d+(?:\.\d+)?)?(?:\s+(?:to|similar\sto|Ma\sto)\s+\d+(?:\.\d+)?|\s*-\s*\d+(?:\.\d+)?)?\s*Ma\b'
    matches = re.findall(pattern, str1)
    return matches


# 3 把excel文件中的age信息全部归到一起 输出日志中 由一个单元格多种形式数据拆分为多行 中心值±+误差两列的形式
# 下一步手动excel分割钱币符号 输出out_all_ages.xlsx文件 带有中心值误差值的表格 仍有多行为同一个矿床名字的那种 下一步要变成一行的（去最大最小 取中位数等措施）
def age_out(file_path):
    # 全局变量 用于控制书写excel的行数
    global row_num1
    row_num1 = 2
    # 创建一个新的Excel工作簿 覆盖all_5w_ages.xlsx
    workbook = openpyxl.Workbook()

    # 创建一个新的工作表
    sheet = workbook.active
    # 设置表头
    sheet['A1'] = 'kc'
    sheet['B1'] = 'cleanen_line'
    sheet['C1'] = 'age'
    sheet['D1'] = 'err'

    op_excel = openpyxl.load_workbook(file_path)
    op_sheet = op_excel['Sheet1']  # 当前工作sheet
    count_row = op_sheet.max_row - 1  # 当前sheet行数
    # print_log = open("result/all_v2.log", 'w')
    # print(count_row)
    for row_index in range(2, count_row + 2):  # 遍历每一行
        if op_sheet.cell(row=row_index, column=3).value != None:
            this_kc = op_sheet.cell(row=row_index, column=2).value
            this_sentence = op_sheet.cell(row=row_index, column=1).value
            original_age = op_sheet.cell(row=row_index, column=3).value

            list_age1 = eval(original_age)  # 识别为列表类型的age
            # print(this_kc,list_age1)
            if (len(list_age1) > 0):
                for age1 in list_age1:
                    # print(age1)
                    if age1 == 'Ma' or age1 == ' Ma' or age1 == ' Ma ':
                        continue
                    else:
                        # age1 = age1.replace(' Ma', '')    #去掉Ma后的时间
                        # # print(kc, wz_num,row_value-1,here_ma,age1,file = print_log)
                        age1 = age1.replace('-', ' - ')
                        age_str_list = age1.split(' ')
                        # print(age_str_list)
                        count_digit = 0
                        out_here_age_num = []
                        for here_split in age_str_list:
                            here_split = here_split.strip()
                            # 1.1 is_number 函数 判断字符串中有几个数字出现 匹配不同策略
                            if is_number(here_split):
                                count_digit += 1  # 看字符串中有几个数字
                                out_here_age_num.append(here_split)
                        # print(count_digit, file = print_log)
                        if count_digit == 0:  # 没有数字(年龄)的
                            continue
                        elif count_digit == 1:  # 只有一个年龄的 直接填充
                            # print('1',out_here_age_num[0],'None',file=print_log)
                            # 2023.9.8 添加下一行 大于4500的扔掉
                            if float(out_here_age_num[0]) < 4500:
                                # 矿床名字 句子 中心值 误差值
                                # print(this_kc+'￥'+this_sentence+'￥' +out_here_age_num[0]+'￥'+'None', file=print_log) # 用excel直接书写代替该行
                                sheet.cell(row=row_num1, column=1,value=this_kc)
                                sheet.cell(row=row_num1, column=2, value=this_sentence)
                                sheet.cell(row=row_num1, column=3,value=out_here_age_num[0])
                                sheet.cell(row=row_num1, column=4,value='None')
                                row_num1 += 1
                                print("已经保存", row_num1-2, '行')
                        else:

                            if age1.find('±') != -1:
                                mid_age1 = out_here_age_num[0]
                                err_age1 = out_here_age_num[1]
                                # 2023.9.8 添加下一行 大于4500的扔掉
                                if float(mid_age1) < 4500:
                                    # 矿床名字 句子 中心值 误差值
                                    # print(this_kc+'￥'+this_sentence+'￥' +mid_age1+'￥'+err_age1, file=print_log)
                                    sheet.cell(row=row_num1, column=1,value=this_kc)
                                    sheet.cell(row=row_num1, column=2, value=this_sentence)
                                    sheet.cell(row=row_num1, column=3,value=mid_age1)
                                    sheet.cell(row=row_num1, column=4,value=err_age1)
                                    row_num1 += 1
                                    print("已经保存", row_num1-2, '行')
                            elif age1.find('to') != -1 or age1.find('-') != -1:
                                mid_age2 = (
                                    float(out_here_age_num[0])+float(out_here_age_num[1]))/2
                                err_age2 = round(abs(float(out_here_age_num[0])-mid_age2), 2)
                                # 2023.9.8 添加下一行 大于4500的扔掉
                                if mid_age2 < 4500:
                                    # 矿床名字 句子 中心值 误差值
                                    # print(this_kc+'￥'+this_sentence+'￥'+str(mid_age2)+'￥'+str(err_age2), file=print_log)
                                    sheet.cell(row=row_num1, column=1,value=this_kc)
                                    sheet.cell(row=row_num1, column=2, value=this_sentence)
                                    sheet.cell(row=row_num1, column=3,value=mid_age2)
                                    sheet.cell(row=row_num1, column=4,value=err_age2)
                                    row_num1 += 1
                                    print("已经保存", row_num1-2, '行')
    # print_log.close()
    # 保存循环之后的excel表格
    # workbook.save('data_source/out_all_ages_new.xlsx') # 没改AE之前的
    workbook.save('final_date/2_twoLine.xlsx')


# 3.1
def is_number(s):
    try:  # 如果能运行float(s)语句，返回True（字符串s是浮点数）
        float(s)
        return True
    except ValueError:  # ValueError为Python的一种标准异常，表示"传入无效的参数"
        pass  # 如果引发了ValueError这种异常，不做任何事情（pass：不做任何事情，一般用做占位语句）
    try:
        import unicodedata  # 处理ASCii码的包
        unicodedata.numeric(s)  # 把一个表示数字的字符串转换为浮点数返回的函数
        return True
    except (TypeError, ValueError):
        pass
    return False


# 4.根据同名数据的正态分布原则去除误差值并保留可信度较高的行 会丢失单行和两行数据差距较大的行修改
'''
1.读取Excel文件。
2.根据名字分组数据。
3.对每个分组进行正态分布异常值检测，保留可信度较高的数据。
4.将处理后的数据输出到新的Excel文件。
'''
def remove_error0():
    import pandas as pd
    import numpy as np
    from scipy import stats

    # 读取Excel文件
    df = pd.read_excel('data_source/test0911.xlsx')
    # 定义阈值（可以根据需要调整） 越小越严格 剩下的越少 2.0表示使用了一个阈值为2.0的标准差 在正态分布中，大约95.4% 的数据点落在均值的两个标准差范围内
    # 原来  8.4w行
    # 散点图 Aktogai矿床为例子  Altar为例子
    # `threshold = 2.0`   只保留大约均值附近95.4%   剩余8w行
    # `threshold = 1.0`  只保留大约均值附近68.26%   剩余7.3w行
    # `threshold = 0.5`  只保留大约均值附近38.3%    剩余5.6w行  较好
    # `threshold = 0.1`  只保留大约均值附近7.96%    剩余1w行
    threshold = 1.0

    # 创建一个新的DataFrame来存储处理后的数据
    cleaned_df = pd.DataFrame(columns=df.columns)

    # 根据名字分组数据并处理每个分组 数据集group
    for name, group in df.groupby('kc'):  # 根据第一列矿床的名字进行分组
        # 计算每个分组的均值和标准差
        # print(type(group['age']))
        # print(group['age'])
        print(len(group['age']))

        mean = group['age'].mean()  # 得到'age'列的均值 mean变量保存这个均值
        std_dev = group['age'].std()  # 保存'age'列的标准差

        # 使用正态分布的方法去除异常值
        # 计算Z分数 Z分数越大，表示数据点偏离均值越远。
        if(len(group['age'])>2):
            z_scores = np.abs(stats.zscore(group['age']))
            valid_rows = (z_scores < threshold)  # 合理行 定义为偏离小于阈值的部分

            # 保留可信度较高的行
            cleaned_group = group[valid_rows]

            # 将处理后的分组数据添加到新的DataFrame中
            cleaned_df = pd.concat([cleaned_df, cleaned_group])
        else:
            z_scores = np.abs(stats.zscore(group['age']))
            print(type(z_scores))
            valid_rows = ((z_scores< threshold) | (z_scores> threshold) | (z_scores is  threshold) )  # 合理行 
            cleaned_group = group[valid_rows]
            cleaned_df = pd.concat([cleaned_df, cleaned_group])
    # 将处理后的数据保存到新的Excel文件
    cleaned_df.to_excel('data_source/test0911_out.xlsx', index=False)


# 4.根据同名数据的正态分布原则去除误差值并保留可信度较高的行
'''
1.读取Excel文件。
2.根据名字分组数据。
3.对每个分组进行正态分布异常值检测，保留可信度较高的数据。
4.将处理后的数据输出到新的Excel文件。
'''
def remove_error1():
    import pandas as pd
    import numpy as np
    from scipy import stats

    # 读取原始Excel文件
    # df = pd.read_excel('data_source/test0911.xlsx')
    # df = pd.read_excel('data_source/out_all_ages_new_tag.xlsx')
    # df = pd.read_excel('data_source/out_all_ages_new.xlsx') # 没改AE之前的
    df = pd.read_excel('final_date/2_twoLine.xlsx')

    # 按照名字分组
    grouped = df.groupby('kc')

    # 创建一个空的DataFrame来存储结果
    filtered_data = pd.DataFrame(columns=df.columns)

    # `threshold = 2.0`   只保留大约均值附近95.4%  
    # `threshold = 1.0`  只保留大约均值附近68.26%   
    # `threshold = 0.5`  只保留大约均值附近38.3%    
    # `threshold = 0.1`  只保留大约均值附近7.96%   
    threshold = 1.0

    # 遍历每个名字的数据
    for name, group_data in grouped:
        if len(group_data) <= 2:  # 如果数据少于等于2个，直接保留
            filtered_data = pd.concat([filtered_data, group_data], ignore_index=True)
        else:
            # 计算数据的均值和标准差
            mean = np.mean(group_data['age'])
            std_dev = np.std(group_data['age'])
            
            # 使用均值和标准差来过滤数据
            filtered_group_data = group_data[abs(group_data['age'] - mean) <= threshold * std_dev] # threshold 数字越小 保留越少
            filtered_data = pd.concat([filtered_data, filtered_group_data], ignore_index=True)

    # 将结果保存到新的Excel文件
    # filtered_data.to_excel('data_source/test0911_out.xlsx', index=False)
    # filtered_data.to_excel('data_source/out_all_ages_new_tag_quxia.xlsx', index=False)
    # filtered_data.to_excel('data_source/out_all_ages_new_quxia1.0.xlsx', index=False) # 没改AE之前的
    filtered_data.to_excel('final_date/3_1_quxia1.0.xlsx', index=False)
    

# 6.假如取 去除瑕疵之后的 平均值的话
def test_mean():
    import pandas as pd

    # 读取原始Excel文件
    df = pd.read_excel('data_source/out_all_ages_new_quxia1.0.xlsx')

    # 使用groupby按照名字分组，并计算每个分组中数字列的平均值
    result_df = df.groupby('kc', as_index=False)['age'].mean() # kc和age分别为第一列和第三列的表格列名

    # 创建一个新的Excel工作簿并将结果写入
    result_df.to_excel('data_source/out_all_ages_new_quxia1.0_mean.xlsx', index=False)


# 7.单独 提取属性1：成矿岩体U-Pb年龄 ,在 out_all_ages_new.xlsx(已经是中心值+偏差的形式)文件的基础上处理。
#  需要注意之后还要去瑕疵和归一两个步骤. 师兄的代码是extract_module.py那个
def tag_extract():
    # 提取属性1：成矿岩体U-Pb年龄
    # 提取规则：
    # （1）出现zircon或者U-Pb或者206Pb/238U；
    # （2）出现ore-bearing/ore-hosting/mineralized/mineralised/mineralization/mineralisation但不出现barren；保留前者，如果只出现barren不保留年龄；如果都不出现保留年龄并放入新属性列‘unknown1’；
    # （3）在单位Ma前，如100+-5 Ma；
    # 在out_all_ages_new.xlsx文件的基础上处理 因为这个表格是提取过的都有Ma关键字的数据了 不用再考虑Ma了,只需要寻找同时出现列表1\2里面的字符串就可以了

    import pandas as pd

    # 读取原始Excel文件
    excel_file_path = 'data_source/out_all_ages_new.xlsx'  # 替换为你的Excel文件路径
    df = pd.read_excel(excel_file_path)

    # 列表1和列表2 从句子里面寻找同时含有这两个列表中的关键字的
    list1 = ['zircon U-Pb', 'Zircon U-Pb', 'zircon', 'Zircon', 'U-Pb', 'P porphyry', '206Pb/238U']     # 属性1 2 3的规则1 # 替换为你的列表1元素
    list2 =  ['ore-bearing', 'ore-hosting', 'ore-forming', 'mineralized', 'mineralised', 'mineralization', 'mineralisation', 'ore is hosted by',
            'host ~ of ore', 'ore-related', 'ore-field', 'Cu(Mo)-bearing', 'Cu-bearing', 'metallogenic age', 'ore formation', 'orebodies']  # 替换为你的列表2元素

    # 处理句子列，将float类型转换为字符串
    df['cleanen_line'] = df['cleanen_line'].astype(str)

    # 匹配含有列表1和列表2中元素的句子，并保留该行
    filtered_df = df[df['cleanen_line'].apply(lambda x: any(word in x for word in list1)) &   # cleanen_line 为excel表格中的句子列的名称
                    df['cleanen_line'].apply(lambda x: any(word in x for word in list2))]

    # 添加匹配到的列表元素
    filtered_df['匹配列表1元素'] = filtered_df['cleanen_line'].apply(lambda x: next(word for word in list1 if word in x))
    filtered_df['匹配列表2元素'] = filtered_df['cleanen_line'].apply(lambda x: next(word for word in list2 if word in x))

    # 保留需要的列
    filtered_df = filtered_df[['kc', 'cleanen_line', '匹配列表1元素', '匹配列表2元素', 'age', 'err']]

    # 保存结果到新的Excel文件
    output_file_path = 'data_source/out_all_ages_new_tag.xlsx'  # 替换为你想要保存的Excel文件路径
    filtered_df.to_excel(output_file_path, index=False)

    print("处理完成，结果已保存到", output_file_path)


# 8.对比pCu数据,将我们的数据的名字全部以列表形式输出出来
# 将一个excel文件中的第一列中的元素按照以下格式进行输出，书写python代码
#['a', 'b', c', 'd']
def pCu():
    import pandas as pd
    # 读取Excel文件，假设第一列名称为 'Column1'
    excel_file_path = 'data_source/out_all_ages_new_quxia1.0_maxMean.xlsx'  # 最终所有归一的Excel文件路径
    # excel_file_path = 'data_source/out_all_ages_new_tag_quxia1_maxMean.xlsx'  # 最终tag属性1归一的Excel文件路径
    df = pd.read_excel(excel_file_path)

    # 提取第一列的元素并按照指定格式输出
    elements = df['kc'].tolist()  # 假设第一列名称为 'kc'
    formatted_output = [f"'{elem.strip()}'" for elem in elements]

    # 将列表转换为字符串并输出
    output_string = '[' + ', '.join(formatted_output) + ']'
    print(output_string)


# 9.运行compare函数跟将最后大的归一数据和pCu数据对比
def compare_one():
    import pandas as pd

    list_ex = ['Accha', 'Adami', 'Agarak', 'Agua Rica', 'Aguablanca', 'Agylki', 'Aikengdelesite', 'Aitik', 'Aketasi', 'Akshatau', 'Aksug', 'Aktogai', 'Alakha', 'Aleksandrovskoe', 'Ali Abad', 'Ali Javad', 'Almaden', 'Almalyk', 'Altaids', 'Altar', 'Altar North', 'Alumbrera', 'Ampucao', 'Amur', 'An', 'Andacollo', 'Anderson Mountain', 'Angelo', 'Anjerd', 'Anjiayingzi', 'Anjishan', 'Anqing', 'Antucoya', 'Aobaotu', 'Aolunhua', 'Aoyoute', 'Aqishan', 'Arasbaran', 'Ardlethan', 'Argentina', 'Arizona', 'Ashele', 'Asia', 'Assarel', 'Associated', 'Atlas', 'Awulale', 'Axi', 'Azerbaijan', 'Babaoshan', 'Babine', 'Babine Lake', 'Bacuri', 'Bada', 'Badaguan', 'Baerzhe', 'Bagdad', 'Baghu', 'Baijintaizi', 'Bainaimiao', 'Bairendaba', 'Bairong', 'Baishan', 'Baishantang', 'Baituyingzi', 'Baixiangshan', 'Baiyanghe', 'Baiyinchagan', 'Baiyinnuoer', 'Baiyun', 'Baizhangyan', 'Baizhangzi', 'Bajiazi', 'Bakoudou', 'Bakyrchik', 'Balcooma', 'Balipo', 'Balkash', 'Balkhash', 'Balong', 'Ban Houayxai', 'Banchang', 'Banduo', 'Bangbule', 'Bangong', 'Bangongco', 'Bangpu', 'Banlashan', 'Banska Stiavnica', 'Banxi', 'Baoanzhai', 'Baogutu', 'Baomai', 'Baoshan', 'Baoyintu', 'Barneys Canyon', 'Barry', 'Bashan', 'Batu Hijau', 'Bayan Obo', 'Bayanbaolege', 'Bayugo', 'Be', 'Beidabate', 'Beina', 'Beiya', 'Bell', 'Belt', 'Benso', 'Berezitovoe', 'Bereznyaki', 'Bethlehem', 'Bi\'r Tawilah', 'Bianjiadayuan', 'Big Bell', 'Bilihe', 'Bilugangan', 'Bingham', 'Bingham Canyon', 'Binghugou', 'Bismark', 'Black Mountain', 'Bolcana', 'Boliden', 'Bolong', 'Bonanza', 'Bondar Hanza', 'Bor', 'Bora', 'Borgulikan', 'Borly', 'Borov Dol', 'Botija', 'Bou Skour', 'Boyongan', 'Bozshakol', 'Brahma', 'Breves', 'British Columbia', 'Broken Hills', 'Buchim', 'Bucim', 'Budunhua', 'Buerkesidai', 'Bujinhei', 'Bulage', 'Bulgaria', 'Butte', 'Bystrinsky', 'Cadia East', 'Cadia Hill', 'Cadia Quarry', 'Caijiaping', 'Caixiashan', 'Calabona', 'Caledonian', 'California', 'Cananea', 'Candelaria', 'Cangyuan', 'Caosiyao', 'Capillitas', 'Caridad', 'Carlton', 'Carmacks', 'Carpathians', 'Casino', 'Caspiche', 'Catface', 'Cavancha', 'Celebration', 'Centralni', 'Cerro Chorcha', 'Cerro Colorado', 'Cerro Corona', 'Cerro Quema', 'Cerro Rico', 'Cerro Verde', 'Cevizlidere', 'Chabu', 'Chagai', 'Chaganbulagen', 'Chaganhua', 'Chagele', 'Chah Firouzeh', 'Chah Zard', 'Chalmers', 'Chalukou', 'Chang\'an', 'Chang\'anbu', 'Chang\'anpu', 'Changfagou', 'Changgou', 'Changhanboluo', 'Changpo', 'Changpu', 'Changpushan', 'Chaobuleng', 'Chapada', 'Charlestown', 'Charlotte', 'Chating', 'Chatree', 'Chehelkureh', 'Chehugou', 'Chelopech', 'Chengba', 'Chengchao', 'Chengmenshan', 'Chentaitun', 'Chigou', 'Chimborazo', 'Chinkuashih', 'Chitudian', 'Chongjiang', 'Chongmuda', 'Choquelimpie', 'Chris', 'Chuchi', 'Chuquicamata', 'Ciemas', 'Cleo', 'Cliff', 'Climax', 'Cobre', 'Cobre Panama', 'Collahuasi', 'Colosa', 'Colquijirca', 'Copler', 'Copperhead', 'Cordilleran', 'Coroccohuayco', 'Coronation Hill', 'Corvet Est', 'Cotabambas', 'Cove', 'Coxheath', 'Crocetta', 'Cuajone', 'Cuihongshan', 'Cuona', 'Cupolymetallic', 'Curich', 'Da Hinggan Mountains', 'Dabaoshan', 'Dabate', 'Dabu', 'Dacaoping', 'Dachang', 'Daero Paulos', 'Dafang', 'Daheishan', 'Dahu', 'Dahutang', 'Dajing', 'Dali', 'Dalingkou', 'Daliugou Formation', 'Dalli', 'Damang', 'Danba', 'Daolundaba', 'Dapai', 'Dapingba', 'Darasun', 'Darreh Zar', 'Darreh Zereshk', 'Daruoluolong', 'Dashigou', 'Dashihe', 'Dashui', 'Dastakert', 'Dasuji', 'Datongkeng', 'Datuanshan', 'Dawan', 'Dawangding', 'Demingding', 'Dengjitun', 'Derni', 'Deva', 'Dexin', 'Dexing', 'Dianfang', 'Diaoquan', 'Dibao', 'Dinkidi', 'Disseminated', 'Disuga', 'Divrigi', 'Diyanqinamu', 'Dizon', 'Dolores', 'Don Javier', 'Dong\'an', 'Dongbulage', 'Dongfengbeishan', 'Dongfengshan', 'Dongga', 'Donggebi', 'Donggou', 'Dongguashan', 'Donggushan', 'Dongji', 'Dongjun', 'Dongleiwan', 'Dongnan', 'Dongshan', 'Dongshanwan', 'Dongshengmiao', 'Dongtian', 'Dongxiang', 'Dongyang', 'Dongyuan', 'Dongzhongla', 'Doupo', 'Douvray', 'Doyon', 'Dry Creek', 'Duhuangling', 'Dujiadongwu', 'Dulong', 'Duobaoshan', 'Duobuza', 'Duolong', 'Duotoushan', 'El Abra', 'El Arco', 'El Durazno', 'El Galeno', 'El Laco', 'El Salvador', 'El Teniente', 'El Valle', 'Elatsite', 'Eliza', 'Elna', 'Emperor', 'Endako', 'Epithermal', 'Epoch', 'Erdaohezi', 'Erdenet', 'Erdenetiin Ovoo', 'Erdenetuin Obo', 'Ergu', 'Ermi', 'Ermiaogou', 'Ernest Henry', 'Ershiyizhan', 'Ertsberg', 'Escondida', 'Escondida Norte', 'Fakos', 'Far Southeast', 'Favona', 'Fazenda Nova', 'Feie\'shan', 'Feitais', 'Fenghuangshan', 'Fengshan', 'Fengshandong', 'Fin', 'Flat', 'Fort Knox', 'Fozichong', 'Francisco', 'Frieda', 'Frontera', 'Fu\'anpu', 'Fujiashan', 'Fujiawu', 'Fukeshan', 'Furtei', 'Fuwan', 'Fuxing', 'Ga\'erqiong', 'Gacun', 'Gaerqin', 'Gaerqiong', 'Gagok', 'Gaijing', 'Galale', 'Galeno', 'Galinge', 'Galore Creek', 'Gangjiang', 'Gangqiongla', 'Gaofeng', 'Gaogangshan', 'Gaojiabang', 'Gaojiashan', 'Gaosongshan', 'Gaoua', 'Gazu', 'Gebunongba', 'Geita Hill', 'Gibraltar', 'Glenburgh', 'Glojeh', 'Golden Cross', 'Golden Mile', 'Golden Sunlight', 'Golpu', 'Gongpoquan', 'Goonumbla', 'Granberg', 'Grasberg', 'Greece', 'Guadalupe', 'Guandaokou Group', 'Guanshan', 'Guaynopita', 'Gudongkeng', 'Guichon Creek', 'Guilingou', 'Guilinzheng', 'Guinaoang', 'Guishan', 'Gushan', 'Gutian', 'Guyana', 'Habo', 'Habo South', 'Hadamengou', 'Hadamiao', 'Haenam', 'Haft Cheshmeh', 'Haftcheshmeh', 'Hahaigang', 'Haib', 'Haigou', 'Haisugou', 'Halasheng', 'Halasu', 'Halongxiuma', 'Handagai', 'Hannan', 'Haobugao', 'Haopinggou', 'Haquira', 'Hardat Tolgoi', 'Hasancelebi', 'Hashitu', 'Hatu', 'Hehuaping', 'Hehuashan', 'Heishantou', 'Hejiangkou', 'Hekanzi', 'Hemlo', 'Henderson', 'Hercynian', 'Hermyingyi', 'Hersai', 'Hetaoping', 'Highland Valley', 'Hills', 'Honggoushan', 'Honghai', 'Hongling', 'Hongqiling', 'Hongshan', 'Hongshanliang', 'Hongshi', 'Hongyuan', 'Hornet', 'Huaixi', 'Huanggang', 'Huanglongpu', 'Huangshandong', 'Huangshaping', 'Huangshuian', 'Huangtun', 'Huangtupo', 'Huangyangshan', 'Huanren', 'Huanuni', 'Huashan', 'Huatong', 'Huayangchuan', 'Hucun', 'Hucunnan', 'Hugo Dummett', 'Huinquintipa', 'Huitongshan', 'Huize', 'Hujiayu', 'Huojihe', 'Huoshenmiao', 'Hutouya', 'Iju', 'Ilarion', 'Iron Cap', 'Iskra', 'Jebel Ohier', 'Jiadanggen', 'Jiagang', 'Jiama', 'Jiamante', 'Jiamantieliek', 'Jiamusi Massif', 'Jiande', 'Jianfengpo', 'Jiangjiatun', 'Jiangligou', 'Jiangnan', 'Jiangshan', 'Jiaojiguan', 'Jiaoxi', 'Jiapigou', 'Jiawula', 'Jiazishan', 'Jidetun', 'Jiepailing', 'Jigongcun', 'Jiguanshan', 'Jiguanshi', 'Jiguanzui', 'Jilongshan', 'Jinchang', 'Jinchanggouliang', 'Jinchangqing', 'Jincheng', 'Jinchuantang', 'Jinding', 'Jinduicheng', 'Jingbian', 'Jinjiling', 'Jinlonggou', 'Jinman', 'Jinsha', 'Jinshan', 'Jintingling', 'Jintonghu', 'Jiru', 'Jitoushan', 'Jiudingshan', 'Jiusangou', 'Joya', 'Junggar', 'Kadjaran', 'Kaerqueka', 'Kafi', 'Kahang', 'Kain', 'Kairagach', 'Kaladaban', 'Kalatage', 'Kalguty', 'Kalinovskoe', 'Kalmakyr', 'Kanggur', 'Karakartal', 'Karamay', 'Karangahake', 'Kartaldag', 'Kassiteres', 'Katbasu', 'Kay Tanda', 'Kayizi', 'Kekekaerde', 'Kekesai', 'Kekura', 'Kelu', 'Kendekeke', 'Kerman', 'Khatsavch', 'Khingan', 'Kidd Creek', 'Kighal', 'Kingking', 'Kirganik', 'Kirovskoye', 'Kiruna', 'Kiseljak', 'Kisladag', 'Knaben', 'Kochbulak', 'Koloula', 'Konevinskoye', 'Kounrad', 'Kristineberg', 'Kuangbei', 'Kubusu', 'Kucing Liar', 'Kuga', 'Kuh Panj', 'Kunshan', 'Kuntabin', 'Kuoerzhenkuola', 'Kuruer', 'Kuthori', 'Laba', 'Ladolam', 'Lailisigao\'er', 'Lailisigaoer', 'Lakange', 'Lalingzaohuo', 'Lamandau', 'Lamasu', 'Lamo', 'Langdu', 'Langlike', 'Langtongmen', 'Lanjia', 'Lanjiagou', 'Lannitang', 'Laochang', 'Laodou', 'Laojiagou', 'Laojiezi', 'Laoliwan', 'Laoshankou', 'Laowan', 'Laowangzhai', 'Laoyaling', 'Lar', 'Laramide', 'Larong', 'Las Picazas', 'Layer', 'Layered', 'Lazurnoe', 'Leimengou', 'Lek', 'Lengshuibeigou', 'Lengshuigou', 'Lengshuikeng', 'Leon', 'Lepanto', 'Leqingla', 'Lermontovskoe', 'Liangshan', 'Lianhuashan', 'Liaoning', 'Liaotun', 'Lietinggang', 'Liguo', 'Lihir', 'Lijiagou', 'Lijiawan', 'Linghou', 'Linglong', 'Lishan', 'Lishanling', 'Liudaowaizi', 'Liuguan', 'Liuhuangshan', 'Liushashan', 'Liushengdian', 'Liyuan', 'Lizhuang', 'Lone Tree', 'Longgen', 'Longjiangting', 'Longmala', 'Longqiao', 'Longshan', 'Longtoushan', 'Lorraine', 'Los Alisos', 'Los Bronces', 'Los Humos', 'Los Pelambres', 'Luanchuan', 'Luanling', 'Luchun', 'Lucy', 'Luerma', 'Lugokanskoe', 'Luming', 'Lunwei', 'Luobodi', 'Luoboling', 'Luobuzhen', 'Luohe', 'Luokuidong', 'Luoyang', 'Luping', 'Machangqing', 'Madeira', 'Madem Lakkos', 'Madneuli', 'Madrid', 'Maevatanana', 'Magdala', 'Magnetite', 'Magushan', 'Maher Abad', 'Makeng', 'Makou', 'Malala', 'Malanjkhand', 'Malartic', 'Malasongduo', 'Mali', 'Malmbjerg', 'Malmyzh', 'Mangzong', 'Mansa Mina', 'Mantos Blancos', 'Maoduan', 'Maozaishan', 'Maozangsi', 'Maqui Maqui', 'Marathon', 'Maratoto', 'Maria', 'Mariquita', 'Maronia', 'Martabe', 'Martha', 'Masjed Daghi', 'Matou', 'Mayo', 'Mayum', 'Medet', 'Meiduk', 'Meiling', 'Meishan', 'Meixian', 'Mengentaolegai', 'Menggongshan', 'Mengxi', 'Mengya\'a', 'Metaliferi Mountains', 'Mexican', 'Michiquillay', 'Miduk', 'Miedzianka', 'Mike', 'Mikheevka', 'Mikheevskoe', 'Mikheevsky', 'Mikheyevsk', 'Milligan', 'Milpillas', 'Mina', 'Mindanao', 'Mingze', 'Minto', 'Mirador', 'Mongolia', 'Mont', 'Montana', 'Monte Negro', 'Moore', 'Morenci', 'Morrison', 'Mosizaote', 'Mount Carlton', 'Mount Charlotte', 'Mount Kare', 'Mount Leyshon', 'Mount Milligan', 'Mount Pleasant', 'Mount Polley', 'Muguayuan', 'Mujicun', 'Muratdere', 'Murgul', 'Mushan', 'N\'yavlenga', 'Nadun', 'Nakhodka', 'Nanmu', 'Nannihu', 'Naoniushan', 'Naozhi', 'Narigongma', 'Nariniya', 'Naruo', 'Narusongduo', 'Nating', 'Neacola', 'Nena', 'Neoarchean', 'Nevada', 'Neves Corvo', 'Newton', 'Newtongmen', 'Niaz', 'Nihe', 'Nimu', 'Nistru', 'Niutougou', 'Nongping', 'Norte', 'Norway', 'Now Chun', 'Nowchun', 'Nucleus', 'Nuevo Chaquiro', 'Nujiang', 'Nuocang', 'Nuri', 'Nurkazgan', 'Nyankanga', 'Obuasi', 'Ohui', 'Ok Tedi', 'Olympias', 'Olympic Dam', 'Omai', 'Opache', 'Oubulage', 'Ouro Roxo', 'Oyu Tolgoi', 'Pakistan', 'Pakistan Chagai', 'Pangjiahe', 'Pangui', 'Panguna', 'Panjia', 'Panormos Bay', 'Paodaoling', 'Paradise Peak', 'Parkam', 'Pasco', 'Pascua', 'Patricia', 'Peak Hill', 'Pebble', 'Pediment', 'Penacho Blanco', 'Perama Hill', 'Perkoa', 'Perseverance', 'Peru', 'Peschanka', 'Petropavlovskoe', 'Philex', 'Phu Kham', 'Pingdingshan', 'Pingshui', 'Pingtan', 'Pingtoubei', 'Pirquitas', 'Plaka', 'Plavica', 'Poboya', 'Poieni', 'Pontides', 'Porgera', 'Portugal', 'Postcollisional', 'Precambrian', 'Productora', 'Prominent Hill', 'Pueblo Viejo', 'Pulang', 'Pusangguo', 'Puziwan', 'Qarachilar', 'Qiagong', 'Qiangdui', 'Qiaomaishan', 'Qiaoxiahala', 'Qibaoshan', 'Qiguling', 'Qimantage', 'Qingcaoshan', 'Qinglong', 'Qiongheba', 'Qiushuwan', 'Qiyugou', 'Quanzigou', 'Quebrada Blanca', 'Quellaveco', 'Questa', 'Qulong', 'Radka', 'Radomiro Tomic', 'Radzimowice', 'Raigan', 'Reagan', 'Reconnaissance', 'Recsk', 'Reko Diq', 'Relin', 'Renison', 'Reshui', 'Resolution', 'Ridgeway', 'Rio Blanco', 'Rodeo', 'Romania', 'Rongga', 'Rongna', 'Rosario', 'Rosia Poieni', 'Ruanjiawan', 'Ruoji', 'Russia', 'Sabagaogoumen', 'Sadaigoumen', 'Saibo', 'Saindak', 'Saishitang', 'Salvadora', 'Sams Creek', 'San Blas', 'San Jorge', 'San Pedro', 'Sancha', 'Sanchakou', 'Sandaozhuang', 'Sandiao', 'Sangan', 'Sangbujiala', 'Sangri', 'Sanjiaowo', 'Sankuanggou', 'Sanshandao', 'Santa Cruz', 'Santa Rita', 'Santa Rosa', 'Sanyuangou', 'Sar Cheshmeh', 'Sarcheshmeh', 'Sari Gunay', 'Sarsuk', 'Sarycheku', 'Sawusi', 'Sebuta', 'Sechangi', 'Sedex', 'Seleteguole', 'Sena', 'Senj', 'Serrinha', 'Several', 'Shadan', 'Shakhtama', 'Shamlugh', 'Shangalon', 'Shangfanggou', 'Shanggusi', 'Shangmushui', 'Shapinggou', 'Shapoling', 'Sharang', 'Shaxi', 'Shedong', 'Shengmikeng', 'Shenshan', 'Shesuo', 'Shihu', 'Shijiawan', 'Shilu', 'Shimadong', 'Shimengou', 'Shimensi', 'Shipingchuan', 'Shiweidong', 'Shiwu', 'Shiyaogou', 'Shizhuyuan', 'Shizilishan', 'Shizishan', 'Shizitou', 'Shotgun', 'Shuangjianzishan', 'Shuangqishan', 'Shuangshan', 'Shuibaiyang', 'Shujiadian', 'Shujigou', 'Shuteen', 'Siah Kamar', 'Sierrita', 'Sigma', 'Sin Quyen', 'Single', 'Sinongduo', 'Siruyidie\'er', 'Sisson Brook', 'Skellefte Group', 'Skouries', 'Snowfield', 'Songnuo', 'Songshugou', 'Songxi', 'Sora', 'South Wales', 'Sovereign', 'Spence', 'Squaw Peak', 'St Demetrios', 'Starra', 'Stawell Wonga', 'Stikine', 'Stypsi', 'Suaqui Verde', 'Such', 'Sukhoi Log', 'Summitville', 'Sungun', 'Sunrise Dam', 'Suoerkuduke', 'Sur', 'Suyunhe', 'Sweden', 'Sweet Home', 'Syama', 'Taibudai', 'Tainskoe', 'Taipingchuan', 'Taipinggou', 'Taipingshan', 'Takht', 'Taking', 'Talatui', 'Taldy Bulak', 'Taldybulak Levoberezhny', 'Talitsa', 'Tallberg', 'Taloveis', 'Tampakan', 'Tangbula', 'Tangjiaping', 'Tangse', 'Tangzhangzi', 'Tanjeel', 'Taocun', 'Taohuazui', 'Taolaituo', 'Taolin', 'Taoxihu', 'Taqian', 'Tarutinsk', 'Tasite', 'Taurus', 'Tawuerbieke', 'Taylor Creek', 'Teniente', 'Tepeoba', 'Tertiary', 'Thanewasna', 'Tharsis', 'Thirteen', 'Tianhexing', 'Tianhu', 'Tianmugou', 'Tiantang', 'Tiantangshan', 'Tiegelong', 'Tiegelongnan', 'Tien Shan', 'Tieshan', 'Tighza', 'Tinggong', 'Tintaya', 'Tl', 'Tocantinzinho', 'Tolgoi', 'Tominsk', 'Tongchang', 'Tongchanggou', 'Tongcun', 'Tonggou', 'Tongguanshan', 'Tongjing', 'Tongkeng', 'Tongkengzhang', 'Tongkuangyu', 'Tonglushan', 'Tongmugou', 'Tongshan', 'Tongshankou', 'Tongshanling', 'Tonian', 'Toquepala', 'Toromocho', 'Tosham', 'Touzha', 'Touzlar', 'Tricolor', 'Tsagaan Suvarga', 'Tuanjiegou', 'Tuketuke', 'Tumen', 'Turkish', 'Turquoise Gulch', 'Turquoise Ridge', 'Tuwu', 'Uchkoshkon', 'Ujina', 'Ulan Uzhur', 'Ulandler', 'Ulsan', 'Unlike', 'Urals', 'Uranium', 'Uspallata', 'Vaikijaur', 'Valea Morii', 'Valley', 'Variscan', 'Veliki Krivelj', 'Verde', 'Victoria', 'Vlaykov Vruh', 'Voluntad', 'Voznesensk', 'Vysokogorskoe', 'Wafi', 'Waiorongomai', 'Waisoi', 'Walegen', 'Wallaby', 'Wanbaoyuan', 'Wandao', 'Wangjiazhuang', 'Wangpingxigou', 'Washan', 'Wassa', 'Weibao', 'Weijia', 'Weilasituo', 'Weiquan', 'Wenquan', 'Weondong', 'Widespread', 'Witwatersrand', 'Wolverine', 'Wonga', 'Woodlawn', 'Woxi', 'Wu', 'Wudaogou', 'Wulaga', 'Wulandele', 'Wulanwuzhuer', 'Wulong', 'Wunugetu', 'Wunugetushan', 'Wurinitu', 'Wushan', 'Wuziqilong', 'Xiadeboli', 'Xiadian', 'Xiajinbao', 'Xianglushan', 'Xiangshan', 'Xiaobaishitou', 'Xiaodonggou', 'Xiaoduobaoshan', 'Xiaofan', 'Xiaohekou', 'Xiaohongshilazi', 'Xiaohulishan', 'Xiaojiashan', 'Xiaojiayingzi', 'Xiaokele', 'Xiaokelehe', 'Xiaoliugou', 'Xiaolonghe', 'Xiaonanshan', 'Xiaorequanzi', 'Xiaosigou', 'Xiaotazigou', 'Xiaotongjiapuzi', 'Xiaotuergen', 'Xiaowolong', 'Xiaoxinancha', 'Xiasai', 'Xiatongling', 'Xiayingfang', 'Xiemisitai', 'Xierqu', 'Xietongmen', 'Xifanping', 'Xigou', 'Xikuangshan', 'Xilekuduke', 'Xiletekehalasu', 'Xiling', 'Xin\'gaguo', 'Xinan', 'Xingjiashan', 'Xingluokeng', 'Xingmeng', 'Xingshan', 'Xinhualong', 'Xinliaodong', 'Xinping', 'Xinqiao', 'Xinxing', 'Xiongcun', 'Xiongmei', 'Xishadegai', 'Xishan', 'Xiuwacu', 'Xoconostle', 'Xuejiping', 'Xulaojiugou', 'Yaguila', 'Yaman Kasy', 'Yanbei', 'Yandong', 'Yangbin', 'Yangchang', 'Yangchongli', 'Yangchuling', 'Yangdong', 'Yanghuidongzi', 'Yangjingou', 'Yangla', 'Yangshan', 'Yanjiagou', 'Yanshan Epoch', 'Yanxi', 'Yao\'an', 'Yaochong', 'Yaojiagou', 'Yaojialing', 'Yashan', 'Yata', 'Yazigou', 'Yechangping', 'Yejiwei', 'Yemaquan', 'Yerington', 'Yili', 'Yindonggou', 'Yingchengzi', 'Yingwuling', 'Yinjiagou', 'Yinkeng', 'Yinshan', 'Yinyan', 'Yixingzhai', 'Yongping', 'Yongxin', 'Youmapo', 'Yu\'erya', 'Yuanzhuding', 'Yubileinoe', 'Yuchiling', 'Yudai', 'Yuejinshan', 'Yueyang', 'Yuhai', 'Yujiadian', 'Yuku', 'Yuleken', 'Yulekenhalasu', 'Yulong', 'Yuntoujie', 'Zafarghand', 'Zalaga', 'Zaldivar', 'Zaozigou', 'Zarshuran', 'Zebuxia', 'Zefreh', 'Zelenodol', 'Zhaceqiao', 'Zhaiwa', 'Zhangsangou', 'Zhaxikang', 'Zhengguang', 'Zhenyuan', 'Zhibula', 'Zhifang', 'Zhifanggou', 'Zhilingtou', 'Zhireken', 'Zhonghe', 'Zhongjia', 'Zhongliao', 'Zhongqiuyang', 'Zhongtiaoshan', 'Zhujiding', 'Zhunsujihua', 'Zhunuo', 'Zhushahong', 'Zhuxi', 'Zhuxiling', 'Zhuyuangou', 'Zijinshan', 'Zoujiashan', 'Zudong']

    # 定义要读写的excel文件路径
    file_path1 = 'data_source/pCu_deps_pros_update.xlsx'
    print_log = open("out/pCu_maxMean.txt",'w')

    df1 = pd.read_excel(file_path1, sheet_name='pCu_deps_pros_update', header=0)


    for kc_name in list_ex:       
        tag_1 = False
        tag_2 = False
        kc_name1 = kc_name.rstrip()
        for index_row in df1.index:
            p_name = df1.loc[index_row].values[4].rstrip()    #原矿床名称
            # p_name = df1.loc[index_row].values[5]  #判断矿床别名是否存在
            p_Ma = df1.loc[index_row].values[15]
            if tag_1 == True and kc_name1 == p_name :       #重复数据
                print(p_name+','+str(p_Ma)+',repeat',file = print_log)  
                tag_2 = True
            if kc_name1 == p_name and tag_2 != True:
                tag_1 = True
                print(p_name+','+str(p_Ma),file = print_log)
            
        if tag_1 != True :
            print(kc_name1+','+'not found',file=print_log)


# 10.运行compare函数跟将提取属性1的tag小数据和pCu数据对比
def compare_tag():
    import pandas as pd
    import openpyxl

    # tag最后归一的数据的所有矿床名字 根据这个去美国地调局查找数据
    list_ex = ['Agua Rica', 'Aguablanca', 'Aikengdelesite', 'Akshatau', 'Aksug', 'Almalyk', 'Altar', 'Alumbrera', 'An', 'Andacollo', 'Anjerd', 'Anjiayingzi', 'Anjishan', 'Ashele', 'Asia', 'Assarel', 'Atlas', 'Axi', 'Badaguan', 'Bainaimiao', 'Bairendaba', 'Bairong', 'Baishan', 'Baituyingzi', 'Baiyanghe', 'Baiyinchagan', 'Baiyinnuoer', 'Baiyun', 'Baizhangyan', 'Bajiazi', 'Bakyrchik', 'Balkash', 'Balkhash', 'Bangpu', 'Banlashan', 'Banxi', 'Baogutu', 'Baoshan', 'Barry', 'Bayan Obo', 'Bayanbaolege', 'Be', 'Beiya', 'Bell', 'Belt', 'Bereznyaki', 'Bianjiadayuan', 'Big Bell', 'Bilihe', 'Bingham', 'Black Mountain', 'Bolong', 'Bonanza', 'Bondar Hanza', 'Bor', 'Borly', 'Botija', 'Bou Skour', 'Bozshakol', 'Breves', 'Bucim', 'Cadia East', 'Caijiaping', 'Caixiashan', 'Caledonian', 'Caosiyao', 'Caridad', 'Carlton', 'Chabu', 'Chaganbulagen', 'Chah Zard', 'Chalukou', 'Chang\'an', 'Changfagou', 'Changpu', 'Charlotte', 'Chating', 'Chelopech', 'Chengmenshan', 'Chentaitun', 'Chigou', 'Chimborazo', 'Chongjiang', 'Chris', 'Chuquicamata', 'Cobre', 'Colosa', 'Coronation Hill', 'Cuajone', 'Cuihongshan', 'Dabaoshan', 'Dabate', 'Dabu', 'Dachang', 'Dafang', 'Daheishan', 'Dahu', 'Dahutang', 'Dajing', 'Daolundaba', 'Dapai', 'Dashui', 'Dasuji', 'Datongkeng', 'Dawangding', 'Dengjitun', 'Dexin', 'Dexing', 'Dong\'an', 'Dongbulage', 'Donggebi', 'Donggou', 'Dongguashan', 'Donggushan', 'Dongshanwan', 'Dongtian', 'Dongyuan', 'Doupo', 'Dry Creek', 'Duhuangling', 'Dulong', 'Duobaoshan', 'Duobuza', 'Duolong', 'El Arco', 'El Salvador', 'El Teniente', 'Elatsite', 'Erdaohezi', 'Erdenet', 'Ergu', 'Ermi', 'Ershiyizhan', 'Escondida', 'Feie\'shan', 'Fenghuangshan', 'Fort Knox', 'Fozichong', 'Francisco', 'Frontera', 'Fujiashan', 'Fujiawu', 'Fukeshan', 'Fuxing', 'Ga\'erqiong', 'Galale', 'Galore Creek', 'Gangjiang', 'Gaogangshan', 'Gaojiabang', 'Gaojiashan', 'Gaosongshan', 'Gazu', 'Gibraltar', 'Glenburgh', 'Golden Mile', 'Golden Sunlight', 'Gongpoquan', 'Goonumbla', 'Grasberg', 'Guilinzheng', 'Guinaoang', 'Gutian', 'Habo', 'Hadamengou', 'Hadamiao', 'Haigou', 'Haisugou', 'Halasheng', 'Handagai', 'Haobugao', 'Hashitu', 'Hehuaping', 'Hehuashan', 'Heishantou', 'Henderson', 'Hercynian', 'Hermyingyi', 'Hills', 'Hongling', 'Hongshan', 'Hongshanliang', 'Hongshi', 'Hongyuan', 'Hornet', 'Huanggang', 'Huanglongpu', 'Huangshandong', 'Huangshaping', 'Huanren', 'Huanuni', 'Huashan', 'Huitongshan', 'Hutouya', 'Iju', 'Jebel Ohier', 'Jiama', 'Jiamante', 'Jiande', 'Jianfengpo', 'Jiangjiatun', 'Jiangnan', 'Jiangshan', 'Jiaoxi', 'Jiapigou', 'Jiawula', 'Jidetun', 'Jiepailing', 'Jiguanshan', 'Jinchang', 'Jinchanggouliang', 'Jinduicheng', 'Jinjiling', 'Jiru', 'Jiusangou', 'Junggar', 'Kadjaran', 'Kaladaban', 'Kalatage', 'Kalinovskoe', 'Kalmakyr', 'Kekekaerde', 'Kelu', 'Kerman', 'Khatsavch', 'Kighal', 'Kiruna', 'Kiseljak', 'Knaben', 'Kounrad', 'Kuangbei', 'Kuh Panj', 'Kuntabin', 'Kuoerzhenkuola', 'Lailisigao\'er', 'Lailisigaoer', 'Lakange', 'Lamasu', 'Laojiagou', 'Laoshankou', 'Laowan', 'Laowangzhai', 'Larong', 'Leimengou', 'Lek', 'Lengshuigou', 'Lengshuikeng', 'Leqingla', 'Lianhuashan', 'Liaoning', 'Lietinggang', 'Lijiagou', 'Linglong', 'Liudaowaizi', 'Liyuan', 'Lizhuang', 'Longgen', 'Longmala', 'Longtoushan', 'Los Bronces', 'Los Humos', 'Los Pelambres', 'Lucy', 'Luerma', 'Luming', 'Lunwei', 'Luoboling', 'Luohe', 'Luoyang', 'Machangqing', 'Maher Abad', 'Makeng', 'Makou', 'Maoduan', 'Maozaishan', 'Maozangsi', 'Mariquita', 'Medet', 'Meiling', 'Meixian', 'Menggongshan', 'Mengxi', 'Mengya\'a', 'Miduk', 'Miedzianka', 'Mikheevka', 'Mina', 'Mongolia', 'Monte Negro', 'Moore', 'Mount Carlton', 'Mount Charlotte', 'Mujicun', 'Murgul', 'Mushan', 'Nadun', 'Nakhodka', 'Nanmu', 'Naoniushan', 'Naozhi', 'Naruo', 'Narusongduo', 'Neoarchean', 'Neves Corvo', 'Newtongmen', 'Nihe', 'Nongping', 'Norte', 'Norway', 'Nujiang', 'Nuri', 'Nurkazgan', 'Olympias', 'Olympic Dam', 'Omai', 'Oyu Tolgoi', 'Panjia', 'Pebble', 'Peru', 'Peschanka', 'Pingtan', 'Plaka', 'Plavica', 'Poieni', 'Precambrian', 'Pueblo Viejo', 'Pulang', 'Qiagong', 'Qibaoshan', 'Qingcaoshan', 'Qinglong', 'Qiongheba', 'Qiushuwan', 'Qiyugou', 'Quanzigou', 'Quellaveco', 'Qulong', 'Relin', 'Resolution', 'Rio Blanco', 'Rongna', 'Ruanjiawan', 'Sadaigoumen', 'Saishitang', 'Salvadora', 'San Pedro', 'Sangan', 'Sankuanggou', 'Sanyuangou', 'Sarcheshmeh', 'Sarsuk', 'Sarycheku', 'Sawusi', 'Several', 'Shadan', 'Shakhtama', 'Shamlugh', 'Shangalon', 'Shangfanggou', 'Shangmushui', 'Shapinggou', 'Sharang', 'Shenshan', 'Shipingchuan', 'Shiweidong', 'Shiyaogou', 'Shizhuyuan', 'Shizitou', 'Shuangjianzishan', 'Shuangqishan', 'Shuangshan', 'Siah Kamar', 'Sin Quyen', 'Skouries', 'Songnuo', 'South Wales', 'Sukhoi Log', 'Suoerkuduke', 'Sweden', 'Taibudai', 'Taipinggou', 'Taking', 'Taldybulak Levoberezhny', 'Tangjiaping', 'Tangzhangzi', 'Taocun', 'Taohuazui', 'Taolaituo', 'Teniente', 'Tertiary', 'Tianhu', 'Tiantangshan', 'Tiegelongnan', 'Tieshan', 'Tinggong', 'Tolgoi', 'Tongchang', 'Tongchanggou', 'Tonggou', 'Tongjing', 'Tongkeng', 'Tongkuangyu', 'Tonglushan', 'Tongshan', 'Tongshankou', 'Tongshanling', 'Toquepala', 'Touzlar', 'Tsagaan Suvarga', 'Tuanjiegou', 'Tuwu', 'Ulandler', 'Unlike', 'Urals', 'Uranium', 'Valley', 'Variscan', 'Wafi', 'Wallaby', 'Washan', 'Wassa', 'Weilasituo', 'Weiquan', 'Wenquan', 'Woxi', 'Wu', 'Wudaogou', 'Wulandele', 'Wulong', 'Wunugetushan', 'Wurinitu', 'Wushan', 'Xiadian', 'Xiajinbao', 'Xianglushan', 'Xiangshan', 'Xiaobaishitou', 'Xiaodonggou', 'Xiaofan', 'Xiaohekou', 'Xiaokele', 'Xiaoliugou', 'Xiaolonghe', 'Xiaoxinancha', 'Xiasai', 'Xiayingfang', 'Xiemisitai', 'Xierqu', 'Xietongmen', 'Xikuangshan', 'Xiletekehalasu', 'Xingluokeng', 'Xinqiao', 'Xiongcun', 'Xiongmei', 'Xishan', 'Xiuwacu', 'Xulaojiugou', 'Yaguila', 'Yandong', 'Yangchang', 'Yanghuidongzi', 'Yangjingou', 'Yangla', 'Yaochong', 'Yaojiagou', 'Yaojialing', 'Yashan', 'Yata', 'Yazigou', 'Yechangping', 'Yejiwei', 'Yerington', 'Yili', 'Yingwuling', 'Yinshan', 'Yixingzhai', 'Yongping', 'Yongxin', 'Yu\'erya', 'Yuchiling', 'Yudai', 'Yueyang', 'Yuhai', 'Yujiadian', 'Yulong', 'Yuntoujie', 'Zaozigou', 'Zarshuran', 'Zebuxia', 'Zefreh', 'Zhaceqiao', 'Zhengguang', 'Zhibula', 'Zhongjia', 'Zhujiding', 'Zhunsujihua', 'Zhunuo', 'Zhuxi', 'Zijinshan', 'Zoujiashan']
    # 定义要读写的excel文件路径
    file_path1 = 'data_source/pCu_deps_pros_update.xlsx'
    print_log = open("out/pCu_tag_maxMean.txt",'w')

    df1 = pd.read_excel(file_path1, sheet_name='pCu_deps_pros_update', header=0)


    for kc_name in list_ex:       
        tag_1 = False
        tag_2 = False
        kc_name1 = kc_name.rstrip()
        for index_row in df1.index:
            p_name = df1.loc[index_row].values[4].rstrip()    #原矿床名称
            # p_name = df1.loc[index_row].values[5]  #判断矿床别名是否存在
            p_Ma = df1.loc[index_row].values[15]
            if tag_1 == True and kc_name1 == p_name :       #重复数据
                print(p_name+','+str(p_Ma)+',repeat',file = print_log)  
                tag_2 = True
            if kc_name1 == p_name and tag_2 != True:
                tag_1 = True
                print(p_name+','+str(p_Ma),file = print_log)
            
        if tag_1 != True :
            print(kc_name1+','+'not found',file=print_log)


if __name__ == '__main__':
    # 未改变AE之前的路径
    # filetxt_path = 'out/all_Ma_out.txt'
    # file1_path = 'data_source/all_5w_ages_new.xlsx'

    filetxt_path = 'out/all_Ma_out_quAE.txt'
    file1_path = 'final_date/1_oneLine.xlsx'

    # 1.调用去重函数 去除txt文件中的重复句子
    # multxt_to_onetxt1()
    # print("已全部去除重复.")


    # 1.1这里补一个步骤  国庆期间 要求AE单独替换成± OCR错误的问题 函数在AE_tran.py


    # 2.去重后的txt提取到第一个三列的表格 第三列为多种数据格式
    # run_all(filetxt_path)  

    # 3.三列未拆分的表格拆分为多行(中心值和误差两列数据的形式) 矿名仍有重复
    # age_out(file1_path)   # 1 第一版的表格 - > 输出多行表格  中心值 误差两列数据形式 pandas版本问题 这个别忘了传入的excel 的第一个表格名字叫Sheet1

    # 4.调用去除误差数据的函数 去除偏差较大的数据 保留较为集中的一部分 正态 
    #   remove_error0()有问题 有的矿床名字下的数据可能距离均值都比较远 都被扔了 还要保留么这种
    #   remove_error1()改进了 没问题了 不丢失
    # remove_error1()
    # print("处理后的数据已经保存到新的Excel文件.")

    # 5.想合适的算法把去掉异常值之后的数据 归一 去最大最小取平均值
    # 这里使用去最大最小取中值的形式 在to_one.py文件中实现

    # 6.假如取 去除瑕疵之后的 平均值的话
    # test_mean()

    # 7.单独 提取属性1：成矿岩体U-Pb年龄 之后再调用4、5函数去除瑕疵然后归一 单独领出去写可以
    # tag_extract()

    # 8.pCu 备用 抽取我们的归一或者tag最终数据的名字一列,为了方便下一步运行compare函数跟pCu数据对比
    # pCu()

    # 9.运行compare函数跟将最后大的归一数据和pCu数据对比
    # compare_one()

    # 10.运行compare函数跟将提取属性1的tag小数据和pCu数据对比
    # compare_tag()















